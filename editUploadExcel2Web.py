import time

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.expected_conditions import visibility_of_element_located
from selenium.webdriver.support.wait import WebDriverWait


driver = webdriver.Chrome()
driver.implicitly_wait(3)

def update_excel(file_path, colRequired, item, new_price):
    lict = {}
    book = openpyxl.load_workbook(file_path)
    sheet = book.active
    for i in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=i).value == colRequired:
            lict["col"] = i

    for r in range(1, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            if sheet.cell(row=r, column=c).value == item:
                lict["row"] = r

    sheet.cell(row=lict["row"], column=lict["col"]).value = new_price
    book.save(r"C:\Users\ASHUTOSH\Downloads\download.xlsx")

driver.get('https://rahulshettyacademy.com/upload-download-test/index.html')

actual_price=int(driver.find_element(By.XPATH, "//div[@id='row-4']/div[4]/div").text)

file_path= r"C:\Users\ASHUTOSH\Downloads\download.xlsx"
item="Kivi"
columnRequired="price"
new_price=28000
update_excel(file_path, columnRequired, item, new_price)
upload_btn= driver.find_element(By.CSS_SELECTOR, "input[type='file']")
upload_btn.send_keys(r"C:\Users\ASHUTOSH\Downloads\download.xlsx")

wait= WebDriverWait(driver, 10)

wait.until(visibility_of_element_located((By.XPATH, "//div[contains(text(),'Updated Excel Data Successfully.')]")))

assert driver.find_element(By.XPATH, "//div[@id='row-4']/div[4]/div").text != actual_price
print( "price changed successfully")


time.sleep(5)
driver.close()