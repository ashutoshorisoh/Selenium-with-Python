import openpyxl
book=openpyxl.load_workbook(r'C:\Users\ASHUTOSH\Documents\python excel.xlsx')
sheet=book.active  # this automatically takes the first active sheet
# or sheer=book["sheet1] #this way we can define sheet name explicitly
#print(sheet.title) #output = "Sheet1"

#print(sheet.max_row)
#print(sheet.max_column)

cell=sheet.cell(row=1, column=2).value="ashutosh" #this won't  change the sheet value, just change here in the terminal, but will hold changed value in entire program
#print(cell)

check=sheet['a2'].value
#print(check) #another simple way of getting value of a particular cell

#print(cell)
maxRow= sheet.max_row
maxColumn=sheet.max_column

for i in range(1, maxRow+1):
    if sheet.cell(row=i, column=1).value=="class":
        for j in range(2, maxColumn + 1):
            print(sheet.cell(row=i, column=j).value)

